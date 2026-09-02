#!/usr/bin/env python3
"""PondCube: messy monthly Excel log -> tidy, AI/API-ready datasets."""
import openpyxl, csv, re

SRC = "/sessions/trusting-awesome-johnson/mnt/uploads/2025.7.31.xlsx"
OUT = "/sessions/trusting-awesome-johnson/mnt/outputs"
YEAR, MONTH, N_DAYS = 2025, 7, 31

LOCATIONS = [
    ("Hatchery","Hatchery","Hatchery"),
    ("Fish Tank 1","Fish Tank 1","Fish Tanks"),
    ("Fish Tank 2","Fish Tank 2","Fish Tanks"),
    ("Fish Tank 3","Fish Tank 3","Fish Tanks"),
    ("Block L","Block L","Block L"),
    ("Water storage","Water Storage","Water Storage"),
    ("Aqua","Aqua","Aqua"),
    ("Pond","Pond","Pond"),
]
PARAMS = [("dissolved_oxygen","mg/L"),("temperature","degC"),("ph","pH"),
          ("ammonia","mg/L"),("nitrate","mg/L"),("nitrite","mg/L")]
REMARK_OFFSET = 6
MEAS = [p for p,_ in PARAMS]

def parse_tank(raw):
    if raw is None: return None,None,None
    if isinstance(raw,(int,float)) and not isinstance(raw,bool): return raw,int(raw),None
    s=str(raw).strip(); m=re.search(r"\d+",s)
    num=int(m.group()) if m else None
    note=None if (m and s==m.group()) else f"non-numeric tank label: {s!r}"
    return s,num,note

def num(v):
    if isinstance(v,bool): return None
    if isinstance(v,(int,float)): return float(v)
    return None

def bcol(day): return 2+(day-1)*7

def layout(ws):
    hdr=[r for r in range(1,ws.max_row+1) if str(ws.cell(r,1).value).strip()=="Tank / Date"]
    marker=next(r for r in range(1,ws.max_row+1)
                if any(str(ws.cell(r,c).value).strip().upper()=="AFTERNOON" for c in range(1,15)))
    return [("morning",5,marker-1),("afternoon",hdr[1]+3,ws.max_row)]

def main():
    wb=openpyxl.load_workbook(SRC,data_only=True)
    long_rows=[]; wide={}; qa=[]; tanks={}
    def wkey_row(loc,zone,tr,tn,date,period):
        return wide.setdefault((loc,tn,date,period),{
            "location":loc,"zone":zone,"tank_id":tn,"tank_id_raw":tr,
            "date":date,"period":period,**{p:None for p in MEAS},"remarks":None})
    for sheet,location,zone in LOCATIONS:
        ws=wb[sheet]
        for period,r0,r1 in layout(ws):
            for r in range(r0,r1+1):
                raw=ws.cell(r,1).value
                if raw is None: continue
                tr,tn,note=parse_tank(raw)
                if note: qa.append({"location":location,"period":period,"row":r,"issue":note})
                tanks.setdefault(location,set()).add(tn)
                for day in range(1,N_DAYS+1):
                    c0=bcol(day)
                    rem=ws.cell(r,c0+REMARK_OFFSET).value
                    rem=str(rem).strip() if isinstance(rem,str) and str(rem).strip() else None
                    date=f"{YEAR:04d}-{MONTH:02d}-{day:02d}"
                    for off,(pk,unit) in enumerate(PARAMS):
                        val=num(ws.cell(r,c0+off).value)
                        if val is None: continue
                        long_rows.append({"location":location,"zone":zone,"tank_id":tn,
                            "tank_id_raw":tr,"date":date,"period":period,
                            "parameter":pk,"value":val,"unit":unit})
                        wkey_row(location,zone,tr,tn,date,period)[pk]=val
                    if rem:
                        wkey_row(location,zone,tr,tn,date,period)["remarks"]=rem
                        long_rows.append({"location":location,"zone":zone,"tank_id":tn,
                            "tank_id_raw":tr,"date":date,"period":period,
                            "parameter":"remark","value":rem,"unit":"text"})
    # tidy long
    lc=["location","zone","tank_id","tank_id_raw","date","period","parameter","value","unit"]
    long_rows.sort(key=lambda d:(d["location"],d["tank_id"] or 0,d["date"],d["period"],d["parameter"]))
    with open(f"{OUT}/pondcube_measurements_long.csv","w",newline="") as f:
        w=csv.DictWriter(f,fieldnames=lc); w.writeheader(); w.writerows(long_rows)
    # wide
    wc=["location","zone","tank_id","tank_id_raw","date","period",*MEAS,"remarks"]
    wl=sorted(wide.values(),key=lambda d:(d["location"],d["tank_id"] or 0,d["date"],d["period"]))
    with open(f"{OUT}/pondcube_observations_wide.csv","w",newline="") as f:
        w=csv.DictWriter(f,fieldnames=wc); w.writeheader(); w.writerows(wl)
    # tanks reference
    with open(f"{OUT}/pondcube_tanks_reference.csv","w",newline="") as f:
        w=csv.writer(f); w.writerow(["location","zone","tank_id"])
        for sheet,location,zone in LOCATIONS:
            for t in sorted(x for x in tanks.get(location,set()) if x is not None):
                w.writerow([location,zone,t])
    # qa
    with open(f"{OUT}/pondcube_data_quality.csv","w",newline="") as f:
        w=csv.DictWriter(f,fieldnames=["location","period","row","issue"]); w.writeheader(); w.writerows(qa)
    meas=sum(1 for r in long_rows if r["parameter"]!="remark")
    print(f"long_total={len(long_rows)} measurements={meas} remarks={len(long_rows)-meas}")
    print(f"wide_rows={len(wl)} distinct_tanks={sum(len({x for x in s if x is not None}) for s in tanks.values())}")
    print(f"QA={qa}")

if __name__=="__main__": main()
