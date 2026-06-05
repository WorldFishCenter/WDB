#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["pandas>=2.0", "openpyxl>=3.1"]
# ///
"""dict_enricher — deterministic tidy-shape gate + value-domain extractor for WDB.

Given one CSV/XLSX, this script:
  1. Validates that the file is ONE tidy table with a single header row. It FAILS
     (exit 2) with a precise list of problems on multi-row/merged headers, metadata
     rows above the header, ragged rows, multiple tables per sheet, or pivot/crosstab
     layouts -- so it doubles as a pre-submit data-validity gate.
  2. Detects the table SHAPE: wide (one row per entity, one column per variable) or
     long (a variable/parameter column + a value column, one row per measurement).
  3. Extracts value domains in a SHAPE-AWARE way -- distinct values for categoricals
     (listed when low-cardinality, else count + examples), ranges for numerics, and for
     long data a range + unit PER parameter (never one mixed domain).

It prints a human-readable report plus a `## Columns`-ready block of domain FACTS for a
maintainer (or the dict-enricher agent) to merge into the matching `_dict.md`. It never
invents prose meaning and never writes to the source file.

Run it (uv supplies pandas; nothing to install):
    uv run .claude/scripts/dict_enricher.py <table.csv|.xlsx> [flags]

Flags:
    --shape {wide,long}     force the shape instead of detecting it
    --var-col NAME          long: column holding the variable/parameter name
    --value-col NAME        long: column holding the measured value
    --unit-col NAME         long: column holding the unit
    --sheet NAME            xlsx: which sheet to read (default: the only data sheet)
    --max-list N            list distinct values when <= N, else count+examples (default 30)
    --examples N            how many examples to show for high-cardinality columns (default 5)
    --json                  emit machine-readable JSON instead of the text report

Exit codes: 0 = valid (domains emitted); 2 = invalid shape (problems listed, no domains);
1 = usage / read error.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

# Column-name hints used for shape detection (case-insensitive, exact match).
VAR_NAME_HINTS = {"parameter", "variable", "var", "measure", "metric", "indicator", "attribute"}
VALUE_NAME_HINTS = {"value", "val", "measurement", "quantity", "result", "reading", "amount"}
UNIT_NAME_HINTS = {"unit", "units", "uom"}

# A unit guessed from a parameter-name suffix when no unit column exists.
SUFFIX_UNITS = {
    "percent": "%", "pct": "%",
    "kcal_kg": "kcal/kg", "mj_kg": "MJ/kg", "kg": "kg", "g": "g", "mg": "mg",
    "mg_l": "mg/L", "mgl": "mg/L", "degc": "degC", "celsius": "degC",
}


class InvalidShape(Exception):
    """Raised with the full list of structural problems found."""

    def __init__(self, problems: list[str]):
        self.problems = problems
        super().__init__("; ".join(problems))


# --------------------------------------------------------------------------- #
# Loading + structural validation
# --------------------------------------------------------------------------- #
def _looks_numeric(s: str) -> bool:
    s = s.strip()
    if not s:
        return False
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        return False


def _looks_date(s: str) -> bool:
    s = s.strip()
    # Conservative: ISO-ish or common date column names used as values.
    return bool(re.fullmatch(r"\d{4}[-/]\d{1,2}([-/]\d{1,2})?", s)) or bool(
        re.fullmatch(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", s)
    )


def _validate_csv_structure(path: Path, max_list: int) -> list[list[str]]:
    """Read the CSV as raw rows and run the structural gate. Returns the rows
    (header + data) if valid; raises InvalidShape otherwise."""
    with path.open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh))
    # Drop wholly-empty trailing rows (trailing newline artifacts).
    while rows and all(c.strip() == "" for c in rows[-1]):
        rows.pop()
    problems: list[str] = []
    if not rows:
        raise InvalidShape(["file is empty"])
    header = rows[0]
    data = rows[1:]

    # --- header cells: present, unique, no merged/multi-row artifacts ---------
    blank_hdr = [i + 1 for i, c in enumerate(header) if c.strip() == ""]
    if blank_hdr:
        problems.append(
            f"blank column header(s) at position(s) {blank_hdr} -- a merged or "
            f"multi-row header, or a missing column name. Give every column one plain name."
        )
    seen: dict[str, int] = {}
    dupes = set()
    for c in header:
        key = c.strip().lower()
        if key:
            seen[key] = seen.get(key, 0) + 1
            if seen[key] > 1:
                dupes.add(c.strip())
    if dupes:
        problems.append(
            f"duplicate column header(s) {sorted(dupes)} -- merged cells or a repeated "
            f"name. Each column needs a unique name."
        )

    # --- title / metadata row above the header --------------------------------
    hdr_nonempty = sum(1 for c in header if c.strip())
    widths = [sum(1 for c in r if c.strip()) for r in data[:8]]
    if widths and hdr_nonempty <= 1 and max(widths) >= 3:
        problems.append(
            "the first row looks like a title/metadata row (1 filled cell) sitting above "
            "the real header. Remove rows above the column headers so row 1 IS the header."
        )

    # --- interior blank rows / multiple tables --------------------------------
    blank_interior = [
        i + 2 for i, r in enumerate(data[:-1]) if all(c.strip() == "" for c in r)
    ]
    if blank_interior:
        problems.append(
            f"blank separator row(s) at line(s) {blank_interior} mid-table -- looks like "
            f"more than one table stacked in the file. Put one table per file."
        )
    hdr_key = tuple(c.strip().lower() for c in header)
    repeat_hdr = [
        i + 2
        for i, r in enumerate(data)
        if tuple(c.strip().lower() for c in r) == hdr_key
    ]
    if repeat_hdr:
        problems.append(
            f"the header row repeats at line(s) {repeat_hdr} -- looks like several tables "
            f"concatenated. Put one table per file."
        )

    # --- ragged rows ----------------------------------------------------------
    ncol = len(header)
    ragged = [
        i + 2
        for i, r in enumerate(data)
        if any(c.strip() for c in r) and len(r) != ncol
    ]
    if ragged:
        shown = ragged[:10]
        more = f" (+{len(ragged) - 10} more)" if len(ragged) > 10 else ""
        problems.append(
            f"ragged row(s) at line(s) {shown}{more}: their column count differs from the "
            f"header's ({ncol}). Every row must have exactly {ncol} columns."
        )

    # --- pivot / crosstab: values used as column names ------------------------
    valueish = [c for c in header if _looks_numeric(c) or _looks_date(c)]
    if len(header) >= 3 and len(valueish) >= max(2, len(header) / 2):
        problems.append(
            f"{len(valueish)} of {len(header)} column names are numbers/dates "
            f"({valueish[:6]}...) -- this looks like a pivot/crosstab. Reshape to long: one "
            f"column naming the variable and one `value` column."
        )

    if problems:
        raise InvalidShape(problems)
    if not data:
        raise InvalidShape(["the file has a header but no data rows."])
    return rows


def _load(path: Path, args) -> "pd.DataFrame":
    import pandas as pd

    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        _validate_csv_structure(path, args.max_list)  # gate first
        sep = "\t" if suffix == ".tsv" else ","
        return pd.read_csv(path, sep=sep, encoding="utf-8-sig")
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return _load_xlsx(path, args)
    raise SystemExit(f"unsupported file type: {suffix} (expected .csv/.tsv/.xlsx)")


def _load_xlsx(path: Path, args) -> "pd.DataFrame":
    import pandas as pd
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    problems: list[str] = []
    nonempty_sheets = [ws.title for ws in wb.worksheets if ws.max_row and ws.max_row > 1]
    if not args.sheet and len(nonempty_sheets) > 1:
        problems.append(
            f"the workbook has {len(nonempty_sheets)} non-empty sheets ({nonempty_sheets}). "
            f"Submit one table per file, or pass --sheet NAME."
        )
    target = args.sheet or (nonempty_sheets[0] if nonempty_sheets else wb.sheetnames[0])
    ws = wb[target] if target in wb.sheetnames else wb.active
    # merged cells almost always mean a multi-row / grouped header
    merged = list(getattr(ws, "merged_cells", []).ranges) if hasattr(ws, "merged_cells") else []
    if merged:
        problems.append(
            f"sheet '{target}' has {len(merged)} merged cell range(s) (e.g. {str(merged[0])}) "
            f"-- merged/multi-row headers aren't allowed. Flatten to one header row."
        )
    if problems:
        raise InvalidShape(problems)
    df = pd.read_excel(path, sheet_name=target)
    # reuse the header checks (blank/duplicate names show up as 'Unnamed: N'/.1 suffixes)
    bad = [c for c in df.columns if str(c).startswith("Unnamed:")]
    if bad:
        raise InvalidShape(
            [f"blank column header(s) {bad} on sheet '{target}' -- merged or missing names."]
        )
    return df


# --------------------------------------------------------------------------- #
# Shape detection
# --------------------------------------------------------------------------- #
def detect_shape(df, args) -> dict:
    cols = list(df.columns)
    lower = {str(c).strip().lower(): c for c in cols}

    def pick(explicit, hints):
        if explicit:
            if explicit in cols:
                return explicit
            if explicit.lower() in lower:
                return lower[explicit.lower()]
            raise SystemExit(f"--*-col '{explicit}' is not a column: {cols}")
        for h in hints:
            if h in lower:
                return lower[h]
        return None

    var_col = pick(args.var_col, VAR_NAME_HINTS)
    value_col = pick(args.value_col, VALUE_NAME_HINTS)
    unit_col = pick(args.unit_col, UNIT_NAME_HINTS)

    if args.shape == "wide":
        return {"shape": "wide", "source": "forced"}
    if args.shape == "long":
        if not (var_col and value_col):
            raise SystemExit(
                "--shape long needs a variable and a value column; pass --var-col / --value-col."
            )
        return {"shape": "long", "source": "forced", "var_col": var_col,
                "value_col": value_col, "unit_col": unit_col}

    # auto-detect
    if var_col and value_col:
        return {"shape": "long", "source": "detected", "var_col": var_col,
                "value_col": value_col, "unit_col": unit_col}
    if bool(var_col) ^ bool(value_col):
        found = var_col or value_col
        raise InvalidShape([
            f"shape is ambiguous: found a '{found}' column but not its partner. For LONG data "
            f"both a variable column ({sorted(VAR_NAME_HINTS)}) and a value column "
            f"({sorted(VALUE_NAME_HINTS)}) are expected. Pass --shape and --var-col/--value-col, "
            f"or use --shape wide."
        ])
    return {"shape": "wide", "source": "detected"}


# --------------------------------------------------------------------------- #
# Domain extraction
# --------------------------------------------------------------------------- #
def _fmt_num(x) -> str:
    f = float(x)
    if f == int(f) and abs(f) < 1e15:
        return str(int(f))
    return f"{f:.6g}"


def _series_kind(s) -> str:
    import pandas as pd
    from pandas.api import types as pt

    name = str(s.name).lower()
    if pt.is_bool_dtype(s):
        return "boolean"
    if pt.is_numeric_dtype(s):
        return "numeric"
    if pt.is_datetime64_any_dtype(s):
        return "datetime"
    if ("date" in name or "time" in name):
        parsed = pd.to_datetime(s, errors="coerce")
        if parsed.notna().mean() >= 0.9:
            return "datetime"
    return "categorical"


def _describe_categorical(s, max_list, n_ex) -> dict:
    vals = s.dropna().astype(str).map(str.strip)
    distinct = sorted(vals.unique().tolist())
    n = len(distinct)
    miss = int(s.isna().sum()) + int((vals == "").sum())
    avg_len = vals.map(len).mean() if len(vals) else 0
    out = {"kind": "categorical", "n_distinct": n, "n_missing": miss}
    if n == 0:
        out["note"] = "empty throughout"
    elif n <= max_list:
        out["values"] = [v for v in distinct if v != ""]
    else:
        # high-cardinality: id-like (every value unique) or free text (long strings)
        kind = "id" if n == len(vals) else ("free_text" if avg_len > 40 else "high_cardinality")
        out["kind"] = kind
        out["examples"] = [v for v in distinct[:n_ex] if v != ""]
    return out


def _describe_numeric(s) -> dict:
    clean = s.dropna()
    out = {"kind": "numeric", "n_missing": int(s.isna().sum())}
    if len(clean):
        out["min"] = _fmt_num(clean.min())
        out["max"] = _fmt_num(clean.max())
    return out


def _describe_datetime(s) -> dict:
    import pandas as pd

    parsed = pd.to_datetime(s, errors="coerce")
    clean = parsed.dropna()
    out = {"kind": "datetime", "n_missing": int(parsed.isna().sum())}
    if len(clean):
        out["min"] = str(clean.min().date())
        out["max"] = str(clean.max().date())
    return out


def describe_column(s, max_list, n_ex) -> dict:
    if s.dropna().empty:
        return {"kind": "empty", "n_missing": int(s.isna().sum())}
    kind = _series_kind(s)
    if kind == "numeric":
        return _describe_numeric(s)
    if kind == "datetime":
        return _describe_datetime(s)
    if kind == "boolean":
        return {"kind": "boolean", "values": sorted(s.dropna().unique().astype(str).tolist()),
                "n_missing": int(s.isna().sum())}
    return _describe_categorical(s, max_list, n_ex)


def _guess_unit_from_name(name: str) -> str | None:
    low = str(name).lower()
    for suf, unit in SUFFIX_UNITS.items():
        if low.endswith("_" + suf) or low.endswith(suf):
            return unit
    return None


def extract_wide(df, max_list, n_ex) -> list[dict]:
    out = []
    for c in df.columns:
        d = describe_column(df[c], max_list, n_ex)
        d["column"] = str(c)
        out.append(d)
    return out


def extract_long(df, shape, max_list, n_ex) -> dict:
    import pandas as pd

    var_col, value_col, unit_col = shape["var_col"], shape["value_col"], shape.get("unit_col")
    dims = [c for c in df.columns if c not in {value_col, unit_col}]
    dim_domains = []
    for c in dims:
        d = describe_column(df[c], max_list, n_ex)
        d["column"] = str(c)
        dim_domains.append(d)

    # value summarized PER parameter -- never one mixed domain
    per_param = []
    vals = pd.to_numeric(df[value_col], errors="coerce")
    non_numeric = int(vals.isna().sum() - df[value_col].isna().sum())
    for param, idx in df.groupby(df[var_col].astype(str)).groups.items():
        v = vals.loc[idx].dropna()
        unit = None
        if unit_col is not None:
            units = df.loc[idx, unit_col].dropna().astype(str).unique().tolist()
            unit = units[0] if len(units) == 1 else (units if units else None)
        if unit is None:
            unit = _guess_unit_from_name(param)
        entry = {"parameter": str(param), "n": int(len(v)), "unit": unit}
        if len(v):
            entry["min"], entry["max"] = _fmt_num(v.min()), _fmt_num(v.max())
        per_param.append(entry)
    per_param.sort(key=lambda e: e["parameter"])
    return {
        "var_col": var_col, "value_col": value_col, "unit_col": unit_col,
        "dimensions": dim_domains, "value_non_numeric": non_numeric, "per_parameter": per_param,
    }


# --------------------------------------------------------------------------- #
# Rendering
# --------------------------------------------------------------------------- #
def _domain_str(d: dict) -> str:
    kind = d["kind"]
    miss = d.get("n_missing", 0)
    miss_s = f" ({miss} missing)" if miss else ""
    if kind == "empty":
        return "empty throughout (no values recorded)"
    if kind == "numeric":
        if "min" in d:
            return f"numeric, range {d['min']}–{d['max']}{miss_s}"
        return f"numeric, all missing"
    if kind == "datetime":
        if "min" in d:
            return f"date, {d['min']} → {d['max']}{miss_s}"
        return "date, all missing"
    if kind == "boolean":
        return "boolean {" + ", ".join(d["values"]) + "}"
    if kind == "categorical":
        if d.get("note"):
            return d["note"]
        if "values" in d:
            return f"{d['n_distinct']} distinct ∈ {{" + ", ".join(d["values"]) + "}" + miss_s
        return f"{d['n_distinct']} distinct"
    # id / free_text / high_cardinality
    label = {"id": "identifier", "free_text": "free text", "high_cardinality": "high-cardinality"}[kind]
    ex = ", ".join(d.get("examples", []))
    return f"{label}, {d['n_distinct']} distinct (e.g. {ex})"


def render_text(report: dict) -> str:
    L = []
    L.append(f"TABLE:   {report['path']}")
    L.append(f"ROWS:    {report['n_rows']:,}    COLUMNS: {report['n_cols']}")
    sh = report["shape"]
    detail = ""
    if sh == "long":
        e = report["long"]
        detail = f"  (var-col={e['var_col']}, value-col={e['value_col']}, unit-col={e['unit_col']})"
    L.append(f"SHAPE:   {sh}  [{report['shape_source']}]{detail}")
    L.append(f"VALIDITY: OK")
    L.append("")
    L.append("## Columns  (value-domain facts — merge into the _dict.md, keep the prose meaning)")
    if sh == "wide":
        for d in report["columns"]:
            L.append(f"- {d['column']}: {_domain_str(d)}")
    else:
        e = report["long"]
        for d in e["dimensions"]:
            L.append(f"- {d['column']}: {_domain_str(d)}")
        L.append(f"- {e['value_col']} (summarize PER {e['var_col']}, never as one domain):")
        for p in e["per_parameter"]:
            unit = p.get("unit")
            unit_s = f" {unit}" if unit and not isinstance(unit, list) else (
                f" units {unit}" if isinstance(unit, list) else " (unit not stated)")
            rng = f"{p['min']}–{p['max']}" if "min" in p else "no numeric values"
            L.append(f"    - {p['parameter']}: {rng}{unit_s}  (n={p['n']})")
        if e["value_non_numeric"]:
            L.append(f"  ⚠ {e['value_non_numeric']} non-numeric value(s) in `{e['value_col']}`.")
    return "\n".join(L)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="Validate a table's shape (tidy wide/long only) and extract value domains.",
    )
    p.add_argument("path", type=Path)
    p.add_argument("--shape", choices=["wide", "long"])
    p.add_argument("--var-col")
    p.add_argument("--value-col")
    p.add_argument("--unit-col")
    p.add_argument("--sheet")
    p.add_argument("--max-list", type=int, default=30)
    p.add_argument("--examples", type=int, default=5)
    p.add_argument("--json", action="store_true")
    args = p.parse_args(argv)

    if not args.path.exists():
        print(f"error: no such file: {args.path}", file=sys.stderr)
        return 1

    try:
        df = _load(args.path, args)
        shape = detect_shape(df, args)
    except InvalidShape as e:
        problems = e.problems
        if args.json:
            print(json.dumps({"path": str(args.path), "valid": False, "problems": problems}, indent=2))
        else:
            print(f"TABLE:   {args.path}")
            print("VALIDITY: FAILED — not a tidy single-table file. Reshape before submitting:\n")
            for i, prob in enumerate(problems, 1):
                print(f"  {i}. {prob}")
            print("\nAllowed shapes: WIDE (one row per entity, one column per variable) or "
                  "LONG (a variable column + a value column).")
        return 2

    report = {
        "path": str(args.path),
        "n_rows": int(len(df)),
        "n_cols": int(len(df.columns)),
        "shape": shape["shape"],
        "shape_source": shape["source"],
        "valid": True,
    }
    if shape["shape"] == "wide":
        report["columns"] = extract_wide(df, args.max_list, args.examples)
    else:
        report["long"] = extract_long(df, shape, args.max_list, args.examples)

    print(json.dumps(report, indent=2) if args.json else render_text(report))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
